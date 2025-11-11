import openpyxl, uuid

path = r"C:\Users\UXK\OneDrive - QlikTech Inc\Documents\Pedro\CCPM\data\bills.xlsx"

wb = openpyxl.load_workbook(path)
ws = wb.active ; ws.title = "debtors"

count = 2

while True:
    value = ws.cell(row=count, column=1).value
    if value is None:
        break
    print(count, value)
    ws.cell(row=count, column=1).value = str(uuid.uuid4())
    count += 1

wb.save(path)
     

arquivo = r"C:\Users\UXK\OneDrive - QlikTech Inc\Documents\Pedro\CCPM\data\bills.xlsx"
wb.save(arquivo)
